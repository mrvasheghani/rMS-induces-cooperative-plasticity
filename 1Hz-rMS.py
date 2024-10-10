# -*- coding: utf-8 -*-
"""
Created on Mon Mar  7 19:41:38 2022

@author: Test-PC
"""
import random
import gc
import os
import numpy as np
import matplotlib.pyplot as plt
from netpyne import specs, sim
from joblib import Parallel, delayed
import multiprocessing
import itertools
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import gc


# Network parameters

netParams = specs.NetParams()       # object of class NetParams to store the network parameters
simConfig = specs.SimConfig()       # object of class SimConfig to store simulation configuration

L23e = {'secs': {}}

L23e['secs']['soma'] = {'geom': {}, 'mechs': {}}
L23e['secs']['soma']['geom'] = {'diam': 96, 'L': 96, 'Ra': 100, 'cm': 1, 'nseg' : 1}
L23e['secs']['soma']['mechs']['pas'] = {'e':-70,'g':0.0001} 
L23e['secs']['soma']['mechs']['hh2'] = {'gnabar': 0.05, 'gkbar': 0.005, 'vtraub':-55 }
L23e['secs']['soma']['mechs']['im'] = {'gkbar':7e-5} 

netParams.cellParams['PYR'] = L23e


## Population parameters
netParams.popParams['S'] = {'cellType': 'PYR', 'numCells': 1}
netParams.popParams['M'] = {'cellType': 'PYR', 'numCells': 1}

## Synaptic mechanism parameters

###############################################################################  Synaptic mechanism parameters
netParams.synMechParams['STD']   = {'mod': 'FDSExp2Syn'} #, 'tau_D1': 140,'d1': 0.7 , 'tau_D2' : 8200}                            # STD synaptic mechanism   

###############################################################################
# Random Stimulation parameters
# simConfig.seeds.conn=np.random.randint(1,256)
simConfig.seeds.stim=np.random.randint(1,256)
# simConfig.seeds.loc=np.random.randint(1,256)

netParams.stimSourceParams['bkg-1'] = {'type': 'NetStim', 'rate': 1, 'noise': 0.5}
netParams.stimTargetParams['bkg->preSy'] = {'source': 'bkg-1', 'sec':'soma', 'loc': 0.5,'conds': {'pop':['S']},'weight': 0.05}

netParams.stimSourceParams['bkg-2'] = {'type': 'NetStim', 'rate': 1, 'noise': 0.5}
netParams.stimTargetParams['bkg->postSy'] = {'source': 'bkg-2', 'sec':'soma', 'loc': 0.5,'conds': {'pop':['M']},'weight': 0.05}
# #############################################################################
###############################################################################
STDPparams = {'STDPon':1, 'hebbwt': 0.00002, 'antiwt':-0.00002, 'wmax': 50, 'tauhebb': 80, 'tauanti':80,'RLon': 0 ,'RLwindhebb': 0, 'useRLexp': 0, 'softthresh': 0, 'verbose':1}
# Cell connectivity rules
netParams.connParams['S->M'] = {'preConds': {'pop': 'S'}, 'postConds': {'pop': ['M']},  #  S -> M
    'probability': 1, #'normal(0.8,0.1)',           # probability of connection
    'weight':  0.01, #'negexp(0.3)',              # synaptic weight
    'delay': 0.5, #1, #'normal(20,8)',                 # transmission delay (ms)
    'sec': 'soma',            # section to connect to
    'loc': 1.0,                 # location of synapse
    'plast': {'mech': 'STDP', 'params': STDPparams},
    'synMech': 'STD'}           # target synaptic mechanism
###############################################################################
# Simulation options


simConfig.duration =    1140*1e3        # Duration of the simulation, in ms  #72*1e3
simConfig.dt = 0.025                # Internal integration timestep to use
simConfig.verbose = True           # Show detailed messages
simConfig.saveCellConns = True
simConfig.recordTraces = {
                          'V_soma':  {'sec': 'soma', 'loc': 0.9, 'var':'v'},
                          'g_STD':   {'sec': 'soma', 'loc': 1.0, 'synMech': 'STD', 'var': 'g'}
                           }
simConfig.hParams = {'celsius': 35, 'v_init': -65.0, 'clamp_resist': 0.001}

simConfig.recordStep = 0.5            # Step size in ms to save data (eg. V traces, LFP, etc)
simConfig.filename = 'test'         # Set file output name
simConfig.savePickle = False        # Save params, network and sim output to pickle file

# simConfig.analysis['plotRaster'] = {'saveFig': True}                    # Plot a raster
simConfig.analysis['plotTraces'] = {'include': [1], 'saveFig': False}  # Plot recorded traces for this list of cells
# simConfig.analysis['plot2Dnet'] = {'saveFig': True}                     # plot 2D cell positions and connections
# simConfig.analysis['plotConn']={'includePre':'S','includePost':'M','feature':'weight', 'saveFig': True }
 
###############################################################################
###############################################################################

threshold=60.00      #0.63 for orginal setting

def Simulation(amp_pre, amp_post):
    simConfig.filename='rMS-1Hz-Continuous-noisy-pre-'+str(amp_pre)+'post(0ms)-'+str(amp_post)
    # COUNT=0

    start_time=   60000
    for ttt in range (0,10000000,10000000):
        for tt in range(0,100000000,100000000):
            for t in range(0,900000,1000):

                StimTime=start_time+t+tt+ttt  
    #             COUNT=COUNT+1
    #             print (StimTime)
    # print ('COUNT='+str(COUNT))
 
                Input_PreSy='Input_PreSy'+ str(ttt)+ str(tt)+str(t) 
                Input_PostSy='Input_PostSy'+ str(ttt)+ str(tt)+str(t)
                Target_PreSy='Target_PreSy'+ str(ttt)+str(tt)+str(t)
                Target_PostSy='Target_PostSy'+ str(ttt)+str(tt)+str(t)
                               
                netParams.stimSourceParams[Input_PreSy] =  {'type': 'IClamp',  'del': StimTime , 'dur': 0.1 , 'amp':threshold*amp_pre/100} 
                netParams.stimSourceParams[Input_PostSy] = {'type': 'IClamp',  'del': StimTime + 0 , 'dur': 0.1 , 'amp':threshold*amp_post/100}
     
                netParams.stimTargetParams[Target_PreSy]  = {'source': Input_PreSy,  'sec':'soma', 'loc': 0.5,'conds': {'pop':['S']}} 
                netParams.stimTargetParams[Target_PostSy] = {'source': Input_PostSy, 'sec':'soma', 'loc': 0.5,'conds': {'pop':['M']}} 
                
    sim.createSimulateAnalyze(netParams = netParams, simConfig = simConfig)
    response = sim.allSimData.g_STD.cell_1
    SimDataTime=sim.simData.t
    return amp_pre, amp_post, SimDataTime, response

num_cores = multiprocessing.cpu_count()-1
amp_pre=range     (137,138,11)
amp_post=range    (125,126,11)
amps = list(itertools.product(amp_pre, amp_post))

chunk_size = 1
chunks = [amps[x:x+chunk_size] for x in range(0, len(amps), chunk_size)]

for chunk in chunks:
    Data=[]
    ModelOutput=Parallel(n_jobs=num_cores)(delayed(Simulation)(amp[0],amp[1]) for amp in chunk)   
    for i in range (chunk_size):
        pre_a=  ModelOutput[i][0]
        post_a= ModelOutput[i][1]
        Time=   ModelOutput[i][2]
        Condactance=ModelOutput[i][3]
        filename='1Hz-pre-'+str(pre_a)+'post-'+str(post_a)
        

        Cond_s=[] 
        Time_s=[]
        BeforStim=[]
        AfterStim=[]

        X=[]
        Y=[]
        Y_Normalized=[]
        
        for j in range (0,len(Time),2000): 
            Cond_max=np.max(Condactance[j:j+2000])
            Time_s.append(round((Time[j]/1000)))          #  conductivity over time (s)
            Cond_s.append(Cond_max)                       #  time (s)
            
        for jj in range (0,len(Time_s),1):                # Prepare for plot
            if Cond_s[jj] > 0.001:
                Y.append(Cond_s[jj])
                X.append(Time_s[jj])
                if jj < 60 :
                    BeforStim.append(Cond_s[jj])
                if jj > 960 :
                    AfterStim.append(Cond_s[jj])      
        Cond_ini = round(np.mean(BeforStim[5:len(BeforStim)]),6)
        Cond_post_norm = round(np.mean(AfterStim[10:len(AfterStim)])/Cond_ini,2) 
############################################################################### plot figures
        plt.figure()
        plt.plot(X, Y,'r.')
        plt.xlabel('Time (seconds)')
        plt.ylabel('Max Conductance')
        plt.title(filename)
        plt.axhline(0,color='gray',linewidth=0.1)
        plt.savefig(filename+'second.png',dpi=600)
    
        Y_Normalized=Y/Cond_ini
        
        plt.figure()
        plt.plot(X,Y_Normalized,'r.')
        plt.xlabel('Time (seconds)')
        plt.ylabel('Normalized Max Conductance')
        plt.title(filename)
        plt.axhline(0, color='gray',linewidth=0.1)
        plt.axhline(1, color='gray',linestyle='--')
        plt.axhline(Cond_post_norm, color='gray',linestyle='dotted',linewidth=0.2)
        plt.text(0,Cond_post_norm , f'{Cond_post_norm}',fontsize=9, ha='right')
        plt.axvline(60, color='gray',linewidth=0.2)
        plt.axvline(960, color='gray',linewidth=0.2)
        plt.savefig(filename+'seconds-Normalized.png',dpi=600)
        
#################################################   Saving Data to excel file
        xlsx_filename = "Sim_Data.xlsx"
        workbook = load_workbook(xlsx_filename) 
        worksheet = workbook.create_sheet(filename)
        worksheet['A1']='Time(second)'
        worksheet['B1']='Conductivity'
        worksheet['C1']='Normalized Conductivity'
        for i, value in enumerate(X, start=2):
            worksheet[f'A{i}']=value
        for i, value in enumerate(Y, start=2):
            worksheet[f'B{i}']=value   
        for i, value in enumerate(Y_Normalized, start=2):
            worksheet[f'C{i}']=value 
        workbook.save(xlsx_filename)
        
        